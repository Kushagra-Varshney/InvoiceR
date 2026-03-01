"""
excel_exporter.py
Builds a styled Excel workbook matching the target layout.

Output format per invoice block:
  Cols 1-9: Invoice Date | Invoice # | Company | Product | Qty | Rate | VAT | SubTotal | Total
  Col 10:   (spacer)
  Cols 11-14: Total Qty | Total VAT | SUB TOTAL | GRAND TOTAL
  Product rows fill cols 1-9; green summary row fills cols 11-14 with totals.
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .aggregator import AggregatedInvoice
from .config import CURRENCY_SYMBOLS, ExcelColors
from .logging_config import get_logger

if TYPE_CHECKING:
    from .invoice_extractor import ExtractionResult

_CLR = ExcelColors()
logger = get_logger("excel")


def _thin_border() -> Border:
    s = Side(style="thin", color="D0D7DE")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


class ExcelExporter:
    """
    Converts a list of AggregatedInvoice objects into a single styled Excel sheet
    matching the target layout with product rows + green summary rows.
    """

    # Column layout — indices are 1-based
    COLS = {
        "invoice_date": 1,
        "invoice_number": 2,
        "company": 3,
        "product_name": 4,
        "quantity": 5,
        "rate": 6,
        "vat": 7,
        "sub_total": 8,
        "total": 9,
        "spacer": 10,  # empty column (matches your sample layout)
        "total_quantity": 11,
        "total_vat": 12,
        "sub_total_sum": 13,
        "grand_total": 14,
    }

    NUM_COLS = 14

    def export(
        self,
        aggregated_invoices: list[AggregatedInvoice],
        raw_results: list[ExtractionResult] | None = None,
    ) -> bytes:
        """
        Build and return the Excel workbook as raw bytes.

        Args:
            aggregated_invoices: List of AggregatedInvoice from Aggregator
            raw_results:         List of ExtractionResult for the Raw Extraction sheet (optional)

        Returns:
            Raw .xlsx bytes ready for download
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"

        self._write_header_row(ws)
        self._set_column_widths(ws)

        current_row = 2

        for inv in aggregated_invoices:
            if not inv.valid:
                current_row = self._write_invalid_row(ws, inv, current_row)
            else:
                current_row = self._write_invoice_block(ws, inv, current_row)

        logger.info(f"Wrote {len(aggregated_invoices)} invoice(s) to Invoices sheet ({current_row - 2} data rows)")

        # Second sheet — raw LLM extraction, no aggregation
        if raw_results:
            self._build_raw_sheet(wb.create_sheet("Raw Extraction"), raw_results)
            logger.info(f"Wrote {len(raw_results)} result(s) to Raw Extraction sheet")

        result_bytes = self._to_bytes(wb)
        logger.info(f"Excel export complete: {len(result_bytes)} bytes")
        return result_bytes

    # ── Sheet builders ────────────────────────────────────────────────────────

    def _write_header_row(self, ws: Worksheet) -> None:
        """Write the column header row matching the target layout."""
        headers = [
            "Invoice Date",
            "Invoice Number",
            "Company",
            "Product Name",
            "Quantity",
            "Price",
            "VAT",
            "Sub Total",
            "Total",
            "",  # spacer
            "Total Quantity",
            "Total VAT",
            "SUB TOTAL",
            "GRAND TOTAL",
        ]

        for col, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.border = _thin_border()
            cell.alignment = _align("center")

            # Yellow for invoice number, blue for the summary columns
            if col == self.COLS["invoice_number"]:
                cell.fill = _fill(_CLR.YELLOW_HDR)
                cell.font = _font(bold=True)
            elif col in (
                self.COLS["total_quantity"],
                self.COLS["total_vat"],
                self.COLS["sub_total_sum"],
                self.COLS["grand_total"],
            ):
                cell.fill = _fill(_CLR.BLUE_HDR)
                cell.font = _font(bold=True)
            else:
                cell.fill = _fill(_CLR.HEADER_BG)
                cell.font = _font(bold=True, color="FFFFFF")

        ws.row_dimensions[1].height = 20

    def _write_invoice_block(self, ws: Worksheet, inv: AggregatedInvoice, start_row: int) -> int:
        """
        Write one invoice block:
          - N product rows (one per product family)
          - 1 green summary row

        Returns the next available row number.
        """
        n_products = len(inv.product_rows)
        summary_row = start_row + n_products  # green row comes after all product rows

        for p_idx, product in enumerate(inv.product_rows):
            row = start_row + p_idx
            is_alt = p_idx % 2 == 1
            row_bg = _CLR.ALT_ROW if is_alt else "FFFFFF"

            # ── Invoice metadata — only on first product row ──────────────────
            if p_idx == 0:
                self._write_cell(
                    ws,
                    row,
                    self.COLS["invoice_date"],
                    inv.document_date,
                    row_bg,
                )
                self._write_cell(
                    ws,
                    row,
                    self.COLS["invoice_number"],
                    inv.document_number,
                    _CLR.INVOICE_NUM_ROW,
                    bold=True,
                )
                self._write_cell(
                    ws,
                    row,
                    self.COLS["company"],
                    inv.vendor_name,
                    row_bg,
                )
            else:
                # Blank metadata cells on subsequent product rows
                for col in (
                    self.COLS["invoice_date"],
                    self.COLS["invoice_number"],
                    self.COLS["company"],
                ):
                    self._write_cell(ws, row, col, "", row_bg)

            # ── Product data ────────────────────────────────────────────
            c = self.COLS
            self._write_cell(ws, row, c["product_name"], product.product_family, row_bg)
            self._write_cell(ws, row, c["quantity"], self._fmt(product.quantity), row_bg, align="right")
            self._write_cell(ws, row, c["rate"], self._to_number(product.unit_price), row_bg, align="right")
            self._write_cell(ws, row, c["vat"], self._fmt(product.vat), row_bg, align="right")
            self._write_cell(ws, row, c["sub_total"], self._fmt(product.sub_total), row_bg, align="right")
            self._write_cell(ws, row, c["total"], self._fmt(product.total), row_bg, align="right")
            self._write_cell(ws, row, self.COLS["spacer"], "", row_bg)

            # Summary columns blank on product rows
            for col in (
                self.COLS["total_quantity"],
                self.COLS["total_vat"],
                self.COLS["sub_total_sum"],
                self.COLS["grand_total"],
            ):
                self._write_cell(ws, row, col, "", row_bg)

        # ── Green summary row ─────────────────────────────────────────────────
        s = inv.summary
        for col in range(1, self.NUM_COLS + 1):
            self._write_cell(ws, summary_row, col, "", _CLR.SUMMARY_BG)

        c = self.COLS
        summary_kw = {"bg_color": _CLR.SUMMARY_BG, "bold": True, "align": "right"}
        self._write_cell(ws, summary_row, c["total_quantity"], self._fmt(s.total_quantity), **summary_kw)
        self._write_cell(ws, summary_row, c["total_vat"], self._fmt(s.total_vat), **summary_kw)
        self._write_cell(ws, summary_row, c["sub_total_sum"], self._fmt(s.sub_total), **summary_kw)
        self._write_cell(ws, summary_row, c["grand_total"], self._fmt(s.grand_total), **summary_kw)

        ws.row_dimensions[summary_row].height = 16

        # One blank spacer row between invoices
        next_row = summary_row + 1
        for col in range(1, self.NUM_COLS + 1):
            ws.cell(row=next_row, column=col).border = _thin_border()

        return next_row + 1

    def _write_invalid_row(self, ws: Worksheet, inv: AggregatedInvoice, row: int) -> int:
        """Write a single red row for invoices that failed validation."""
        self._write_cell(ws, row, self.COLS["invoice_date"], inv.document_date, _CLR.INVALID_BG)
        self._write_cell(ws, row, self.COLS["invoice_number"], inv.document_number, _CLR.INVALID_BG)
        self._write_cell(ws, row, self.COLS["company"], inv.vendor_name, _CLR.INVALID_BG)
        self._write_cell(ws, row, self.COLS["product_name"], f"⚠ {inv.warning}", _CLR.INVALID_BG)

        for col in range(5, self.NUM_COLS + 1):
            self._write_cell(ws, row, col, "", _CLR.INVALID_BG)

        return row + 2  # blank row after invalid

    # ── Shared helpers ────────────────────────────────────────────────────────

    def _write_cell(self, ws, row, col, value, bg_color, bold=False, align="left") -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = _fill(bg_color)
        cell.font = _font(bold=bold)
        cell.border = _thin_border()
        cell.alignment = _align(align, "center")

    def _fmt(self, value: float) -> float:
        """Round to 2 decimal places for display."""
        return round(value, 2)

    def _to_number(self, value: str) -> float | str:
        """Convert a string value to float if possible, stripping currency symbols and commas."""
        if not value:
            return ""
        cleaned = str(value).replace(",", "").replace(" ", "")
        for sym in CURRENCY_SYMBOLS:
            cleaned = cleaned.replace(sym, "")
        cleaned = cleaned.strip()
        try:
            return float(cleaned)
        except ValueError:
            return value  # keep original if not parseable

    def _set_column_widths(self, ws: Worksheet) -> None:
        widths = {
            1: 14,  # Invoice Date
            2: 16,  # Invoice Number
            3: 28,  # Company
            4: 28,  # Product Name
            5: 10,  # Quantity
            6: 12,  # Rate
            7: 14,  # VAT
            8: 16,  # Sub Total
            9: 16,  # Total
            10: 4,  # Spacer
            11: 14,  # Total Quantity
            12: 14,  # Total VAT
            13: 16,  # SUB TOTAL
            14: 16,  # GRAND TOTAL
        }
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

    def _build_raw_sheet(self, ws: Worksheet, results: list[ExtractionResult]) -> None:
        """
        Raw extraction sheet — one row per line item, no grouping or summing.
        Invoices are separated by a blue header row showing invoice metadata.

        Columns:
          File | Page | Doc # | Date | Vendor | Client |
          Stock Code | Supp Code | Product Family | Description |
          Qty | Unit Price | VAT | Line Total
        """
        RAW_HEADERS = [
            "File",
            "Page",
            "Document #",
            "Date",
            "Vendor",
            "Client",
            "Stock Code",
            "Supp Code",
            "Product Family",
            "Description",
            "Quantity",
            "Unit Price",
            "VAT Amount",
            "Line Total",
        ]

        # Header row
        for col, label in enumerate(RAW_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = _fill(_CLR.HEADER_BG)
            cell.font = _font(bold=True, color="FFFFFF")
            cell.border = _thin_border()
            cell.alignment = _align("center")
        ws.row_dimensions[1].height = 20

        # Column widths for raw sheet
        raw_widths = {
            1: 22,
            2: 6,
            3: 16,
            4: 14,
            5: 28,
            6: 28,
            7: 12,
            8: 12,
            9: 20,
            10: 40,
            11: 10,
            12: 12,
            13: 14,
            14: 14,
        }
        for col, width in raw_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        current_row = 2

        for result in results:
            if not result.success:
                # Single red row for failed extractions
                self._write_cell(ws, current_row, 1, result.source_filename, _CLR.INVALID_BG)
                self._write_cell(ws, current_row, 2, result.page_number, _CLR.INVALID_BG)
                self._write_cell(ws, current_row, 3, f"FAILED: {result.error}", _CLR.INVALID_BG)
                for col in range(4, len(RAW_HEADERS) + 1):
                    self._write_cell(ws, current_row, col, "", _CLR.INVALID_BG)
                current_row += 2
                continue

            d = result.data
            line_items = d.line_items or []

            # Blue separator row — invoice metadata
            meta_label = (
                f"  Invoice: {d.document_number or '—'}  |  "
                f"Date: {d.document_date or '—'}  |  "
                f"Vendor: {d.vendor_name or '—'}  |  "
                f"Client: {d.client_name or '—'}  |  "
                f"File: {result.source_filename}  pg.{result.page_number}"
            )
            for col in range(1, len(RAW_HEADERS) + 1):
                cell = ws.cell(row=current_row, column=col, value=meta_label if col == 1 else "")
                cell.fill = _fill(_CLR.SEPARATOR_BG)
                cell.font = _font(bold=True)
                cell.border = _thin_border()
                cell.alignment = _align("left")
            # Merge separator across all columns
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(RAW_HEADERS))
            ws.row_dimensions[current_row].height = 16
            current_row += 1

            if not line_items:
                self._write_cell(ws, current_row, 1, result.source_filename, _CLR.NO_ITEMS_BG)
                self._write_cell(ws, current_row, 3, "No line items extracted", _CLR.NO_ITEMS_BG)
                for col in [2] + list(range(4, len(RAW_HEADERS) + 1)):
                    self._write_cell(ws, current_row, col, "", _CLR.NO_ITEMS_BG)
                current_row += 2
                continue

            for i, item in enumerate(line_items):
                row_bg = _CLR.ALT_ROW if i % 2 == 1 else "FFFFFF"
                values = [
                    result.source_filename,
                    result.page_number,
                    d.document_number or "",
                    d.document_date or "",
                    d.vendor_name or "",
                    d.client_name or "",
                    item.stock_code or "",
                    item.supp_code or "",
                    item.product_family or "",
                    item.description or "",
                    self._to_number(item.quantity or ""),
                    self._to_number(item.unit_price or ""),
                    self._to_number(item.vat_amount or ""),
                    self._to_number(item.line_total or ""),
                ]
                for col, value in enumerate(values, start=1):
                    align = "right" if col >= 11 else "left"
                    self._write_cell(ws, current_row, col, value, row_bg, align=align)
                current_row += 1

            # Blank spacer row between invoices
            for col in range(1, len(RAW_HEADERS) + 1):
                ws.cell(row=current_row, column=col).border = _thin_border()
            current_row += 1

    def _to_bytes(self, wb: openpyxl.Workbook) -> bytes:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
