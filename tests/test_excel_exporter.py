"""Tests for the Excel exporter."""

import io

import openpyxl
import pytest

from invoice_extraction.aggregator import AggregatedInvoice, InvoiceSummary, ProductRow
from invoice_extraction.excel_exporter import ExcelExporter


@pytest.fixture
def exporter():
    return ExcelExporter()


@pytest.fixture
def valid_invoice():
    return AggregatedInvoice(
        source_filename="test.pdf",
        page_number=1,
        document_date="2026-01-15",
        document_number="INV-001",
        vendor_name="Test Vendor",
        client_name="Test Client",
        product_rows=[
            ProductRow(
                product_family="PRODUCT A",
                quantity=10.0,
                unit_price="100.00",
                vat=150.0,
                sub_total=850.0,
                total=1000.0,
            ),
            ProductRow(
                product_family="PRODUCT B",
                quantity=5.0,
                unit_price="200.00",
                vat=150.0,
                sub_total=850.0,
                total=1000.0,
            ),
        ],
        summary=InvoiceSummary(
            total_quantity=15.0,
            total_vat=300.0,
            sub_total=1700.0,
            grand_total=2000.0,
        ),
        valid=True,
    )


@pytest.fixture
def invalid_invoice():
    return AggregatedInvoice(
        source_filename="bad.pdf",
        page_number=1,
        document_number="INV-002",
        vendor_name="Bad Vendor",
        valid=False,
        warning="Cannot parse quantity",
    )


class TestExcelExporter:
    def test_export_produces_bytes(self, exporter, valid_invoice):
        result = exporter.export([valid_invoice])
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_export_valid_xlsx(self, exporter, valid_invoice):
        result = exporter.export([valid_invoice])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Invoices" in wb.sheetnames

    def test_header_row(self, exporter, valid_invoice):
        result = exporter.export([valid_invoice])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Invoices"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 15)]
        assert "Invoice Date" in headers
        assert "Invoice Number" in headers
        assert "GRAND TOTAL" in headers

    def test_product_rows_written(self, exporter, valid_invoice):
        result = exporter.export([valid_invoice])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Invoices"]
        # Row 2 = first product, Row 3 = second product, Row 4 = summary
        assert ws.cell(row=2, column=4).value == "PRODUCT A"
        assert ws.cell(row=3, column=4).value == "PRODUCT B"

    def test_summary_row_totals(self, exporter, valid_invoice):
        result = exporter.export([valid_invoice])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Invoices"]
        # Summary row is after the 2 product rows = row 4
        assert ws.cell(row=4, column=11).value == 15.0  # total_quantity
        assert ws.cell(row=4, column=14).value == 2000.0  # grand_total

    def test_invalid_invoice_row(self, exporter, invalid_invoice):
        result = exporter.export([invalid_invoice])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Invoices"]
        # Invalid invoice should have warning in product_name column
        cell_value = ws.cell(row=2, column=4).value
        assert "Cannot parse quantity" in cell_value

    def test_raw_sheet_created(self, exporter, valid_invoice, sample_extraction_result):
        result = exporter.export([valid_invoice], raw_results=[sample_extraction_result])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Raw Extraction" in wb.sheetnames

    def test_raw_sheet_not_created_without_results(self, exporter, valid_invoice):
        result = exporter.export([valid_invoice])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Raw Extraction" not in wb.sheetnames

    def test_multiple_invoices(self, exporter, valid_invoice, invalid_invoice):
        result = exporter.export([valid_invoice, invalid_invoice])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Invoices"]
        # Should have data for both invoices
        assert ws.max_row > 4

    def test_empty_list(self, exporter):
        result = exporter.export([])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Invoices"]
        # Only header row
        assert ws.max_row == 1
